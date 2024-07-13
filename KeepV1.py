import tkinter as tk
from tkinter import messagebox, ttk, font, filedialog
import openpyxl
import keepaUtils as kp
import time
import threading

def update_gui(tokens, total_productos, productos_procesados, refillIn):
    tokens_var.set(f"Tokens disponibles: {tokens}")
    productos_var.set(f"Productos procesados: {productos_procesados}/{total_productos}")
    refill_var.set(f"Tiempo de espera: {refillIn / 60000:.2f} minutos")

def run_process():  
    def import_excel()->openpyxl.Workbook:            
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            try:
                # Cargar el archivo Excel
                wb = openpyxl.load_workbook(filename)    
                # sheet = wb.active
                # for row in sheet.iter_rows(min_row=2, values_only=True, max_row=3):  # Empezar desde la segunda fila (después de los encabezados)
                #     productos.append(row[0])  
                # consulta_thread = threading.Thread(target=run_process)
                # consulta_thread.start()
            except Exception as e:
                print(f"Error al cargar el archivo Excel: {e}")    
            else:
                return wb
    workbook = import_excel()    
    sheet = workbook.active    
    asins_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Empezar desde la segunda fila (después de los encabezados)        
        asins_list.append(row[0])      
    batch_size = 100
    total_productos = len(asins_list)
    productos_procesados = 0
    wb2 = kp.generarExcel()
    tokens_left,_ = kp.TokenStatus()    
    if tokens_left > 0:
        if tokens_left > batch_size:
            for i in range(0, total_productos, batch_size):
                tokens_left,sleepTime = kp.TokenStatus()
                while tokens_left < batch_size:
                    update_gui(tokens_left, total_productos, productos_procesados, sleepTime)
                    time.sleep(sleepTime//1000)
                    tokens_left,sleepTime = kp.TokenStatus()
                
                batch = asins_list[i:i+batch_size]
                kp.agregarProductosExcel(wb2, kp.RequestProducts(batch))
                tokens_left -= batch_size
                productos_procesados += batch_size
                update_gui(tokens_left, total_productos, productos_procesados, sleepTime)
                if tokens_left < batch_size:
                    update_gui(tokens_left, total_productos, productos_procesados, sleepTime)
                    time.sleep(sleepTime//1000)
                    tokens_left,sleepTime = kp.TokenStatus()     
    kp.guardarExcel(wb2)                           

# def import_excel():        
#     global productos
#     filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
#     if filename:
#         try:
#             # Cargar el archivo Excel
#             wb = openpyxl.load_workbook(filename)    
#             sheet = wb.active
#             for row in sheet.iter_rows(min_row=2, values_only=True, max_row=3):  # Empezar desde la segunda fila (después de los encabezados)
#                 productos.append(row[0])  
#             consulta_thread = threading.Thread(target=run_process)
#             consulta_thread.start()
#         except Exception as e:
#             print(f"Error al cargar el archivo Excel: {e}")              

root = tk.Tk()
root.title("Keepa")
# configurando tamaño
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = screen_width // 8
window_height = screen_height // 8
position_x = screen_width // 2
position_y = screen_height // 4
root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
# Header
header = tk.Frame(root,bg="#0F0F0F")
header.columnconfigure(0, weight=10)
header.rowconfigure(0,weight=10)

# content1 = tk.Frame(root,bg='#949E9B')
# content1.columnconfigure(0, weight=10)
# content1.rowconfigure(0, weight=10)

content2 = tk.Frame(root,bg='#232D3F')
content2.columnconfigure(0, weight=10)
content2.rowconfigure(0, weight=40)
content2.rowconfigure(1, weight=20)
content2.rowconfigure(2, weight=20)
content2.rowconfigure(3, weight=20)

footer = tk.Frame(root,bg='#0F0F0F')
# configuracion columna
root.columnconfigure(0, weight=10)
# configuracion filas
root.rowconfigure(0, weight=1) #10%
root.rowconfigure(1, weight=8) #80%
root.rowconfigure(2, weight=1) #10%
# agregar header, content y footer
header.grid(row=0, sticky=tk.NSEW)
# content1.grid(row=1, column=0, sticky=tk.NSEW)
content2.grid(row=1, sticky=tk.NSEW, columnspan=2)
footer.grid(row=2, sticky=tk.NSEW)

button = tk.Button(content2, text="Importar Excel", command=run_process, justify="center", highlightthickness=2, bg="#008170", activebackground='#005B41', width=20, height=2)
button.grid(column=0, row=0)

tokens_var = tk.StringVar()
productos_var = tk.StringVar()
refill_var = tk.StringVar()

tokens_label = tk.Label(content2, textvariable=tokens_var,bg="#008170")
productos_label = tk.Label(content2, textvariable=productos_var,bg="#008170")
refill_label = tk.Label(content2, textvariable=refill_var,bg="#008170")

tokens_label.grid(column=0 ,row=1)
productos_label.grid(column=0 ,row=2)
refill_label.grid(column=0 ,row=3)
update_gui(0, 0, 0, 0)

root.mainloop()

if __name__ == '__main__':
    pass