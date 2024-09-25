import math
import keepa
import requests
import openpyxl
import numpy as np
import configparser
import pandas as pd
from collections import defaultdict
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

def import_excel(filename)->openpyxl.Workbook:            
    asins_l = []
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook(filename)    
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Empezar desde la segunda fila (después de los encabezados)
            asins_l.append(row[0])  
    except Exception as e:
        print(f"Error al cargar el archivo Excel: {e}")    
    else:
        asins_l = [x for x in asins_l if x is not None]
        return asins_l
            
def getConfig()->configparser.ConfigParser:
    config = configparser.ConfigParser()
    try:
        config.read('config.ini')
        # url_base = config['keepa']['url']
        # api_key = config['keepa']['api_key']
        # domain = config['keepa']['domain']
        # asin = 'B0CFV4WLY6'
    except Exception as e:
        print(e)
        return None
    else:
        return config

def SaveBestSellerList(asins_l, category, domain, date):
    wb = openpyxl.Workbook()
    ws = wb.active
    dom = ''
    if domain == 1:
        dom = 'USA'
    elif domain == 6:
        dom = 'CANADA'
    ws.title = f'BestSeller_ASINS_{dom}_{category}'
    ws.append(['ASINS'])
    for key,value in enumerate(asins_l, start=2):
        ws.cell(row=key, column=1, value=value)
        # ws.append(value)
    wb.save(f'BestSeller_{dom}_{date}.xlsx')   
    
# Función para manejar la serialización de datetime
def json_serial(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()  # Convertir datetime a una cadena ISO 8601
    if isinstance(obj, np.ndarray):
        return obj.tolist()  # Convertir numpy.ndarray a lista
    if isinstance(obj, pd.DataFrame):
        return obj.to_dict(orient='records')  # Convertir DataFrame a lista de diccionarios
    raise TypeError(f"Tipo {type(obj)} no serializable a JSON")

def get_avg(price_list:list, date_list:list, month:int=0, year:int=0):
    precios_month_year = []
    precios_sinNan = [precio for precio in price_list if precio is not None and not math.isnan(precio)]
    lowest = min(precios_sinNan)
    precio_mas_reciente = precios_sinNan[-1]
    fecha_mas_reciente = date_list[-1]
    dt_30_dias = fecha_mas_reciente - timedelta(days=30)
    dt_60_dias = fecha_mas_reciente - timedelta(days=60)
    dt_90_dias = fecha_mas_reciente - timedelta(days=90)
    dt_180_dias =  fecha_mas_reciente - timedelta(days=180)
    
    precios_ultimos_30_dias = [precio for precio, fecha in zip(price_list, date_list) if (fecha >= dt_30_dias) and precio is not None and not math.isnan(precio)]
    precios_ultimos_60_dias = [precio for precio, fecha in zip(price_list, date_list) if (fecha >= dt_60_dias) and precio is not None and not math.isnan(precio)]
    precios_ultimos_90_dias = [precio for precio, fecha in zip(price_list, date_list) if (fecha >= dt_90_dias) and precio is not None and not math.isnan(precio)]
    precios_ultimos_180_dias = [precio for precio, fecha in zip(price_list, date_list) if (fecha >= dt_180_dias) and precio is not None and not math.isnan(precio)]
    
    if month > 0 or year > 0:
        precios_month_year = [precio for precio, fecha in zip(price_list, date_list) if (fecha.month == month and fecha.year == year) and precio is not None and not math.isnan(precio)]

    if precios_ultimos_30_dias:
        promedio_precio_30_dias = sum(precios_ultimos_30_dias) / len(precios_ultimos_30_dias)
    else:
        promedio_precio_30_dias = None

    if precios_ultimos_60_dias:
        promedio_precio_60_dias = sum(precios_ultimos_60_dias) / len(precios_ultimos_60_dias)
    else:
        promedio_precio_60_dias = None
    
    if precios_ultimos_90_dias:
        promedio_precio_90_dias = sum(precios_ultimos_90_dias) / len(precios_ultimos_90_dias)
    else:
        promedio_precio_90_dias = None

    if precios_ultimos_180_dias:
        promedio_precio_180_dias = sum(precios_ultimos_180_dias) / len(precios_ultimos_180_dias)
    else:
        promedio_precio_180_dias = None
    
    if precios_month_year:
        promedio_month_year = sum(precios_month_year) / len(precios_month_year)
    else:
        promedio_month_year = None
    
    # precios_sinNan = [precio for precio in price_list if precio is not None and not math.isnan(precio)]
    # suma = sum(precios_sinNan)
    # tamaño = len(price_list)
    # print(f"totalList:{suma}, tamañoList:{tamaño}, promedio365:{suma/tamaño}")
    promedio_precio_365_dias = sum(precios_sinNan) / len(price_list)

    return(precio_mas_reciente, promedio_precio_30_dias, promedio_precio_60_dias,
           promedio_precio_90_dias, promedio_precio_180_dias, promedio_precio_365_dias, lowest, promedio_month_year)

def getMonthList()->list:
    fecha_mas_reciente = datetime.now()
    meses_del_año = [(fecha_mas_reciente.year, fecha_mas_reciente.month)]
    for _ in range(12):
        fecha_mas_reciente -= relativedelta(months=1)  # Usamos relativedelta para restar meses
        meses_del_año.append((fecha_mas_reciente.year, fecha_mas_reciente.month))
    return meses_del_año

def getAvgMontly(price_list:list, date_list:list)->dict:
    meses_del_año = getMonthList()
    # print(f"meses del año: {meses_del_año}")
    precios_por_mes = defaultdict(list)

    # Recorrer las listas y agrupar por año y mes
    for precio, fecha in zip(price_list, date_list):
        clave_mes_año = (fecha.year, fecha.month)
        precios_por_mes[clave_mes_año].append(precio)

    # Calcular el promedio por mes, si no hay precios, se asigna promedio 0
    promedios_por_mes = {}
    for clave_mes_año in meses_del_año:
        precios = precios_por_mes.get(clave_mes_año, [])
        if precios:
            precios_ = [x for x in precios if x is not None and not math.isnan(x)]
            if len(precios_) > 0:
                promedio = sum(precios_) / len(precios_)
            else:
                promedio = None
        else:
            promedio = None  # Mes sin precios, promedio 0
        promedios_por_mes[clave_mes_año] = promedio
    # print(f"precios por mes: {precios_por_mes}")
    # print(f"promedios por mes: {promedios_por_mes}")
    return promedios_por_mes
    # Mostrar resultados
    # for (año, mes), promedio in sorted(promedios_por_mes.items()):
    #     print(f"Promedio de {mes}/{año}: {promedio:.2f}")

def getDias()->list:
    fecha_actual = datetime.now()
    fecha_hace_un_ano = fecha_actual - relativedelta(years=1)
    lista_fechas = []
    dia = fecha_actual
    while dia >= fecha_hace_un_ano:
        lista_fechas.append(dia)
        dia -= timedelta(days=1)
    return lista_fechas

def PrecioXdia(price_list:list, date_list:list)->list:
    lista_fechas = getDias()
    precio_por_fecha = {fecha.date(): precio for fecha, precio in zip(date_list, price_list)}
    lista_precios = []
    for fecha in lista_fechas:
        precio = precio_por_fecha.get(fecha.date(), None)  # Obtener el precio o 0 si no está
        lista_precios.append(precio)
    return lista_precios

def process_products(asin_list:list, month:int=0, year:int=0)->dict:      
    # columnas_excel = ["asin","domainId","imagesCSV","title","monthlySold","new_current","amazon_current","fbafees","packageWeight",
    #                   "referralFeePercent","hazardousMaterials","new_offer_count_current","lowest_fba_seller","manufacturer","brand",
    #                   "current_salesRanks","30_salesRanks","180_salesRanks","90_salesRanks","180_salesRanks","365_salesRanks",
    #                   "30_new","60_new","90_new","180_new","365_new","30_amazon","60_amazon","90_amazon","180_amazon","365_amazon",
    #                   "buybox_current","30_buybox","60_buybox","90_buybox","180_buybox","365_buybox","upcList","categories"]    
    asins_response = {} 
    precios_dia_new = {}
    precios_dia_amazon = {}
    for i,product in enumerate(asin_list):
        print(f"{product['asin']}_producto_{i+1}")   
        upcList = ""        
        categories = "" 
        current_fba = fba_30 = fba_60 = fba_90 = fba_180 = fba365 = 0
        current_salesRanks = salesRanks_30 = salesRanks_60 = salesRanks_90 = salesRanks_180 = salesRanks_365 = 0  
        new_current = new_30 =  new_60 = new_90 = new_180 = new_365 = 0
        amazon_current = amazon_30 = amazon_60 = amazon_90 = amazon_180 = amazon_365 = 0
        buybox_current = buybox_30 = buybox_60 = buybox_90 = buybox_180 = buybox_365 = 0
        new_offer_count_current = lowest_fba_seller = monthlySold = fbafees = 0
        referralFeePercent = lowest_amazon = lowest_new = mes_avg_amazon = mes_avg_new = 0
        salesRankDrops30 = salesRankDrops90 = salesRankDrops180 = salesRankDrops365 = 0
        lista_precios_dia_new = lista_precios_dia_amazon = []
        avg_m_amazon = {}
        avg_m_new = {}
        if 'data' in product:
            data = product['data']
            if 'NEW' in data:
                if product['data']['NEW'] is not None:
                    new = product['data']['NEW']
                    new_time = product['data']['NEW_time']
                    if not all(np.isnan(new)):
                        avg_m_new = getAvgMontly(new, new_time)
                        lista_precios_dia_new = PrecioXdia(new, new_time)                        
                        lista_precios_dia_new.insert(0,product['asin'])
                        if year > 0 or month > 0:
                            new_current,new_30, new_60, new_90, new_180, new_365, lowest_new, mes_avg_new = get_avg(new, new_time, month=month, year=year)
                        else:
                            new_current,new_30, new_60, new_90, new_180, new_365, lowest_new,_ = get_avg(new, new_time)
            if 'AMAZON' in data:
                if product['data']['AMAZON'] is not None:
                    amazon = product['data']['AMAZON']
                    amazon_time = product['data']['AMAZON_time']
                    if not all(np.isnan(amazon)):
                        avg_m_amazon = getAvgMontly(amazon, amazon_time)
                        lista_precios_dia_amazon = PrecioXdia(amazon, amazon_time)
                        lista_precios_dia_amazon.insert(0,product['asin'])
                        if year > 0 or month > 0:
                            amazon_current,amazon_30, amazon_60, amazon_90, amazon_180, amazon_365, lowest_amazon, mes_avg_amazon = get_avg(amazon, amazon_time, month=month, year=year)
                        else:
                            amazon_current,amazon_30, amazon_60, amazon_90, amazon_180, amazon_365, lowest_amazon,_ = get_avg(amazon, amazon_time)
            if 'BUY_BOX_SHIPPING' in data:
                if product['data']['BUY_BOX_SHIPPING'] is not None:
                    buybox = product['data']['BUY_BOX_SHIPPING']
                    buybox_time = product['data']['BUY_BOX_SHIPPING_time']
                    if not all(np.isnan(buybox)):
                        buybox_current,buybox_30, buybox_60, buybox_90, buybox_180, buybox_365,_,_ = get_avg(buybox, buybox_time)
            if 'SALES' in data:
                if product['data']['SALES'] is not None:
                    sales = product['data']['SALES']
                    sales_time = product['data']['SALES_time']
                    if not all(np.isnan(sales)):
                        current_salesRanks, salesRanks_30, salesRanks_60, salesRanks_90, salesRanks_180, salesRanks_365,_,_ = get_avg(sales, sales_time)
            if 'NEW_FBA' in data:
                if product['data']['NEW_FBA'] is not None:
                    fba = product['data']['NEW_FBA']
                    fba_time = product['data']['NEW_FBA_time']
                    if not all(np.isnan(fba)):
                        current_fba, fba_30, fba_60, fba_90, fba_180, fba365, _,_ = get_avg(fba, fba_time)

        if 'upcList' in product:
            if product['upcList'] is not None:
                upcList = ','.join(product['upcList'])
        if 'categories' in product:
            cate = product['categories']
            if cate is not None:
                categorias_str = [str(categoria) for categoria in product['categories']]
                categories = ','.join(categorias_str)

        asin = product['asin']
        domainId = product['domainId']            
        imagesCSV = product['imagesCSV']
        title = product['title']           
        if 'monthlySold' in product:
            monthlySold = product['monthlySold']
        if 'fbaFees' in product:
            if product['fbaFees'] is not None:
                if product['fbaFees']['pickAndPackFee'] is not None:
                    fbafees = product['fbaFees']['pickAndPackFee']/100
        packageWeight = product['packageWeight']/453.59290944
        if 'referralFeePercent' in product:
            referralFeePercent = product['referralFeePercent']

        hazardousMaterials = ''
        if 'hazardousMaterials' in product:
            hazardousMaterials_list = product['hazardousMaterials']
            aspect_dict = {}            
            for hazard in hazardousMaterials_list:
                aspect = hazard['aspect']
                value = hazard['value']
                if aspect in aspect_dict:
                    aspect_dict[aspect].append(value)
                else:
                    aspect_dict[aspect] = [value]
            
            hazardousMaterials = '; '.join([f"{aspect}: {','.join(values)}" for aspect, values in aspect_dict.items()])            
        manufacturer = product['manufacturer']
        brand = product['brand']   

        if 'stats' in product:
            stats = product['stats'] 
            if stats is not None:
                salesRankDrops30 = product['stats']['salesRankDrops30']
                salesRankDrops90 = product['stats']['salesRankDrops90']
                salesRankDrops180 = product['stats']['salesRankDrops180']
                salesRankDrops365 = product['stats']['salesRankDrops365']
        
        values = [asin,domainId, imagesCSV, title, monthlySold, new_current, amazon_current, fbafees, packageWeight, referralFeePercent, hazardousMaterials,
                    new_offer_count_current, lowest_fba_seller, manufacturer, brand, current_salesRanks, salesRanks_30, salesRanks_60, salesRanks_90, salesRanks_180,
                    salesRanks_365, new_30, new_60, new_90, new_180, new_365, amazon_30, amazon_60, amazon_90, amazon_180, amazon_365, buybox_current,
                    buybox_30, buybox_60, buybox_90, buybox_180, buybox_365, upcList, categories,lowest_amazon,lowest_new,current_fba, fba_30, fba_60, fba_90, fba_180,
                    fba365, salesRankDrops30, salesRankDrops90, salesRankDrops180, salesRankDrops365]
        
        for (año, mes), promedio in avg_m_new.items():
            values.append(promedio)
        
        for (año, mes), promedio in avg_m_amazon.items():
            values.append(promedio)

        if (mes_avg_amazon != 0) or (mes_avg_new != 0):
            values.append(mes_avg_amazon)
            values.append(mes_avg_new)

        asins_response[asin] = values
        precios_dia_new[asin] = lista_precios_dia_new
        precios_dia_amazon[asin] = lista_precios_dia_amazon
        
    return (asins_response, precios_dia_new, precios_dia_amazon)

def generarExcel(category:str, domain:str, n_columna:str)->openpyxl.Workbook:
    columnas_excel = ["asin","domainId","imagesCSV","title","monthlySold","new_current","amazon_current","fbafees","packageWeight",
                      "referralFeePercent","hazardousMaterials","new_offer_count_current","lowest_fba_seller","manufacturer","brand",
                      "current_salesRanks","30_salesRanks","180_salesRanks","90_salesRanks","180_salesRanks","365_salesRanks",
                      "30_new","60_new","90_new","180_new","365_new","30_amazon","60_amazon","90_amazon","180_amazon","365_amazon",
                      "buybox_current","30_buybox","60_buybox","90_buybox","180_buybox","365_buybox","upcList","categories", "lowest_AMAZON", "lowest_NEW",
                      'current_fba', 'fba_30', 'fba_60', 'fba_90', 'fba_180','fba365',"salesRankDrops30","salesRankDrops90","salesRankDrops180","salesRankDrops365"]
    
    col_meses = getMonthList()
    for mes in col_meses:
        columnas_excel.append(f"NEW_{mes[1]}-{mes[0]}")
        # print(f"NEW_{mes[1]}-{mes[0]}")
    
    for mes in col_meses:
        columnas_excel.append(f"AMAZON_{mes[1]}-{mes[0]}")
        # print(f"AMAZON_{mes[1]}-{mes[0]}")

    if n_columna != '':
        columnas_excel.append(n_columna+'_amazon')
        columnas_excel.append(n_columna+'_new')

    dom = ''
    if domain == '1':
        dom = 'US'
    elif domain == '6':
        dom = 'CA'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{dom}_{category}'
    ws.append(columnas_excel)

    string_date_list = [fecha.strftime("%Y-%m-%d") for fecha in getDias()]
    lista_dias = string_date_list
    lista_dias.insert(0, 'ASINS')

    ws_new = wb.create_sheet(title='precios_new')
    ws_new.append(lista_dias)
    ws_amazon = wb.create_sheet(title='precios_amazon')
    ws_amazon.append(lista_dias)
    wb.active = ws
    # print(f"Columnas {columnas_excel}")
    return wb

def agregarProductosExcel(wb:openpyxl.Workbook, productos:dict, new_dia:dict, amazon_dia:dict) -> openpyxl.Workbook:
    ws = wb.active
    for key, value in productos.items():
        ws.append(value)

    
    wb.active = wb["precios_new"]
    ws_new = wb.active
    
    for key, value in new_dia.items():
        ws_new.append(value)
    
    wb.active = wb["precios_amazon"]
    ws_amazon = wb.active
    
    for key, value in amazon_dia.items():
        ws_amazon.append(value)  

    wb.active = ws

def guardarExcel(wb:openpyxl.Workbook, fname):
    now = datetime.now() # current date and time
    date_time = now.strftime("%m%d%Y%H%M%S")
    filename = f'{fname}_{date_time}.xlsx' 
    wb.save(filename)  
    print(f"saved {filename}")     

def guardarBestSeller(filename:str,prods:list, title:str):
    columnas_excel = ['ASINS']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(columnas_excel)
    for idx, valor in enumerate(prods, start=2):
        ws.cell(row=idx, column=1, value=valor)
    wb.save(filename)  
    print(f"saved {filename}")    

def BestSellers(domain:str, category:str, month:str, year:str)->list:
    asins_l = []
    config = getConfig()
    url_base = config['keepa']['url']
    api_key = config['keepa']['api_key']
    api = keepa.Keepa(api_key)
    tokens_left = api.tokens_left
    while tokens_left < 50:
        api.wait_for_tokens()
        tokens_left = api.tokens_left
    print(f"Buscando BestSellers...Tokens:{api.tokens_left}")
    # url = f"{url_base}/bestsellers?key={api_key}&domain={domain}&category={category}&range={range}&month={mont}&year={year}"
    url = f"{url_base}/bestsellers?key={api_key}&domain={domain}&category={category}&month={month}&year={year}"
    payload = {}
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        response_json = response.json()
        if 'bestSellersList' in response_json:
            if response_json['bestSellersList'] is not None:
                asins_l = response_json['bestSellersList']['asinList']
    print(f"Asins_Bestsellers: {len(asins_l)}")
    guardarBestSeller(f"BSellers-{month}-{year}.xlsx", asins_l, f"domain-{domain} Category-{category}")
    # asins_l = ['B0BZ3LPV6J','B000PEOMC8']
    # asins_l = ['B0009H5BLM']
    return asins_l

if __name__ == '__main__':
    pass