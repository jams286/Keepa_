import requests
import configparser
import openpyxl
import csv
from datetime import datetime

json_keys =  [
    "asin", "domainId", "imagesCSV", "title", "monthlySold", "csv.[1].[1]", "csv.[0].[1]",
    "fbafees.pickAndPackFee", "packageWeight", "referralFeePercent", "hazardousMaterials",
    "csv.[11].[1]", "csv.[10].[1]", "manufacturer", "brand"
]

def GetDateTimeFromKeepTime(keepaTime:int):              
    ts = (keepaTime+21564000)*60     
    return datetime.fromtimestamp(ts)

def getConfig()->configparser.ConfigParser:
    config = configparser.ConfigParser()
    try:
        config.read('config.ini')
        # url_base = config['keepa']['url']
        # api_key = config['keepa']['api_key']
        # domain = config['keepa']['domain']
        # asin = 'B0CFV4WLY6'
    except Exception as e:
        print(e)
        return None
    else:
        return config
    
def saveConfig(config:configparser.ConfigParser)->bool:    
    try:
        with open('config.ini','w') as configFile:
            config.write(configFile)
    except Exception as e:
        print(e)
        return False
    else:
        return True
    
def load_excel_data(filename)-> openpyxl.Workbook:
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook(filename)       
    except Exception as e:
        print(f"Error al cargar el archivo Excel: {e}")
    else:
        return wb
    
def load_csv_data(filename)-> openpyxl.Workbook:
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(filename) as f:
            reader = csv.reader(f, delimiter=',')
            for row_index, row in enumerate(reader, start=1):
                for column_index, cell_value in enumerate(row, start=1):
                    ws.cell(row=row_index, column=column_index).value=cell_value    
                       
    except Exception as e:
        print(f"Error al cargar el archivo csv: {e}")
    else:
        return wb

def generarExcel()->openpyxl.Workbook:
    columnas_excel = ["asin","domainId","imagesCSV","title","monthlySold","new_current","amazon_current","fbafees","packageWeight",
                      "referralFeePercent","hazardousMaterials","new_offer_count_current","lowest_fba_seller","manufacturer","brand"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'productos'
    ws.append(columnas_excel)
    return wb
    # wb.save("productos.xlsx")

def agregarProductosExcel(wb:openpyxl.Workbook, productos:dict) -> openpyxl.Workbook:
    ws = wb.active
    for key, value in productos.items():
        ws.append(value)

def guardarExcel(wb:openpyxl.Workbook):
    wb.save('productos.xlsx')

def TokenStatus()->int:    
    config = getConfig()    
    url_base = config['keepa']['url']
    api_key = config['keepa']['api_key']    
    url = f"{url_base}/token?key={api_key}"
    payload = {}
    headers = {}        
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        response_json = response.json()
        tokensLeft = response_json['tokensLeft']
        refillIn = response_json['refillIn']
        return (tokensLeft,refillIn)
    else:
        return 0

def RequestProducts(asin_list:list)->dict:      
    # columnas_excel =  [
    #                 "asin", "domainId", "imagesCSV", "title", "monthlySold", "csv.[1].[1]", "csv.[0].[1]",
    #                 "fbafees.pickAndPackFee", "packageWeight", "referralFeePercent", "hazardousMaterials",
    #                 "csv.[11].[1]", "csv.[10].[1]", "manufacturer", "brand"
    #              ]
    config = getConfig()    
    url_base = config['keepa']['url']
    api_key = config['keepa']['api_key']
    asins = ','.join(asin_list)  
    url = f"{url_base}/product?key={api_key}&domain=1&days={365}&asin={asins}"
    payload = {}
    headers = {}
    asins_response = {}    
    
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        response_json = response.json()        
        for product in response_json['products']:            
            new_current = 0
            amazon_current = 0
            new_offer_count_current = 0
            lowest_fba_seller = 0
            monthlySold = 0
            fbafees = 0
            referralFeePercent = 0
            asin = product['asin']
            domainId = product['domainId']            
            imagesCSV = product['imagesCSV']
            title = product['title']           
            if 'monthlySold' in product:
                monthlySold = product['monthlySold']
            if product['csv'][1] is not None:
                new_current = product['csv'][1][1]/100
            if product['csv'][0] is not None:
                amazon_current = product['csv'][0][1]/100
            if 'fbaFees' in product:
                if product['fbaFees'] is not None:
                    if product['fbaFees']['pickAndPackFee'] is not None:
                        fbafees = product['fbaFees']['pickAndPackFee']/100
            packageWeight = product['packageWeight']/453.59290944
            if 'referralFeePercent' in product:
                referralFeePercent = product['referralFeePercent']
            csv_values = product.get('csv',[])            
            hazardousMaterials = ''
            if 'hazardousMaterials' in product:
                hazardousMaterials_list = product['hazardousMaterials']
                aspect_dict = {}            
                for hazard in hazardousMaterials_list:
                    aspect = hazard['aspect']
                    value = hazard['value']
                    if aspect in aspect_dict:
                        aspect_dict[aspect].append(value)
                    else:
                        aspect_dict[aspect] = [value]
                
                hazardousMaterials = '; '.join([f"{aspect}: {','.join(values)}" for aspect, values in aspect_dict.items()])            
            if product['csv'][11] is not None:
                new_offer_count_current = product['csv'][11][1]/100 
            if product['csv'][10] is not None:
                lowest_fba_seller = product['csv'][10][1]/100
            manufacturer = product['manufacturer']
            brand = product['brand']
            values = [asin,domainId, imagesCSV, title, monthlySold, new_current, amazon_current, fbafees, packageWeight, referralFeePercent, hazardousMaterials,
                      new_offer_count_current, lowest_fba_seller, manufacturer, brand]
            asins_response[asin] = values
    else:
        print(f"Error: {response.status_code}")
    return asins_response

if __name__ == '__main__':
    # pass   
    # print(GetDateTimeFromKeepTime(6628900))
    # from datetime import datetime
    import time
    a = 21564000
    b = 60
    c = 7166604
    ts = (c+a)*b   
    print(datetime.fromtimestamp(ts))
